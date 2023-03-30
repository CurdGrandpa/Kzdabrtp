def test_export_product():
    import openpyxl
    book = openpyxl.Workbook()
    sheet = book.active
    i = j = 1
    for var in vars(Products).keys():
        sheet.cell(row=i, column=j, value=var)
        j += 1
    myData = Products.objects.all()
    for product in myData:
        i += 1
        j = 1
        for var in vars(product).values():
            sheet.cell(row=i, column=j, value=var)
            j += 1

    book.save("ProductsData.xlsx")


def import_product(file_name):
    from openpyxl import load_workbook
    book = load_workbook(file_name)
    try:
        sheet = book.get_sheet_by_name('Sheet')
    except:
        print('Ошибка! Необходим рабочий лист "Sheet"!')
    print(sheet.max_row)
    print(sheet.max_column)
    for i in range(sheet.max_row):
        for j in range(sheet.max_column):
            print(sheet.cell(row=i+1, column=j+1).value)


if __name__ == '__main__':
    export_product()
