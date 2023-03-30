def rowreader(row, keys, sheet):
    i = 0
    _row = dict()
    for key in keys:
        i += 1
        _row[key] = sheet.cell(row=row, column=i).value
    return _row


def import_products():
    # 0-Артикул, 1-Категория, 2-Подкатегория, 3-Подкатегория категории, 4-Номенклатура с упаковкой,
    # 5-Текстовое описание номенклатуры, 6-Имя файла картинки, 7-цена, 8-валюта

    from openpyxl import load_workbook
    book = load_workbook(file_name)
    try:
        sheet = book.get_sheet_by_name('Sheet')
    except:
        print('Ошибка! Необходим рабочий лист "Sheet"!')

    keys = ['inner_code', 'category', 'categories', 'type', 'name', 'description', 'image', 'price', 'currency']

    for i in range(2, sheet.max_row):
        row = rowreader(i, keys, sheet)
        try:                                                                        # Если цена указана
            price = float(row[7].replace(" ", ""))                                  # то он её считывает
        except:                                                                     # Если не указана, то ловит ошибку
            print(f'no price: line {line_count}')
            price = 0                                                               # и ставит цену, равную нулю

        if price <= 0:                                                               # Если цена указана (если товар существует. Ну, или строка заполнена, там)
            continue

        try:
            product = Product.objects.get(inner_code=row[0])        # Чекает, есть ли уже такой продукт (продукт с таким внутренним номером)
        except:                                                     # Если продукта с таким внутренним номером нет
            product = Product(inner_code=row[0])                    # Создаётся новый

        try:
            if not product.block_image_update:							#Если Изменение фотографии при апдейте не заблокированно
                img_name = row['image'].split("/")[-1]						#Берём название картинки
                saved_image = save_img(img_name, row[1])					#Функция, добавляющая изображение, насколько я понял
                if saved_image:
                    product.image = saved_image
                    product.hide_product = False
                else:
                    product.hide_product = True
            product.price = price
            product.current_price = product.count_current_price()
            if not product.sku:
                product.sku = row['inner_code']
            # product.url = row['inner_code']   # Заполняется само, если оставить пустым
            if row[5] != "NULL":
                product.description = row['description']
            product.category = Category.objects.get(name=row['category'])  # Вот здесь осторожно
            product.unit = Unit.objects.get(value=row['name'].split(', ')[1])
            product.name = row['name'].split(', ')[0]
            product.title = product.name
            product.save()
        except:
            print(product.name)
        print(f'Processed {line_count} lines.')                                             # Это полезно

        return "OK"                                                                         # Это приятно

# Это всё надо доработать
# Надо бы добавить сборщик мусора. Чтобы мусор собирал и на сервере проблем не было.
#Буду работать над этим

